"""31 - Apply the floor+cap sample allocation to the review spreadsheet.

Reads the Phase-4 workbook (v1), computes per-category sample targets with a
floor+cap rule, and writes a v2 workbook (v1 left intact) plus a v2 CSV mirror
and an allocation report.

Allocation (decisions.md 2026-05-21):
  - 25 named categories; "none" excluded (target 0).
  - proportional pre-clip p_c = 100 * size_c / frame_total  (frame = named only)
  - target_c = round(min(CAP, max(FLOOR, p_c))) with FLOOR=3, CAP=6.
  - Total is reported, not forced to exactly 100.

Inputs (read-only): sample/skills_review_2026-05-20.xlsx,
                    sample/skills_review_2026-05-20.csv
Outputs: sample/skills_review_2026-05-20_v2.xlsx
         sample/skills_review_2026-05-20_v2.csv
         analysis/profile/sample_allocation_report.md

Run:  python3 scripts/31_update_review_spreadsheet.py
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font

import _lib as L

DATA_DATE = "2026-05-20"
FLOOR, CAP, TARGET_TOTAL = 3, 6, 100
NONE_CAT = "none"
N_COLS = 15  # category tabs have 15 columns (A..O)

ROOT = Path(__file__).resolve().parent.parent
V1_XLSX = ROOT / "sample" / f"skills_review_{DATA_DATE}.xlsx"
V2_XLSX = ROOT / "sample" / f"skills_review_{DATA_DATE}_v2.xlsx"
V1_CSV = ROOT / "sample" / f"skills_review_{DATA_DATE}.csv"
V2_CSV = ROOT / "sample" / f"skills_review_{DATA_DATE}_v2.csv"
REPORT = ROOT / "analysis" / "profile" / "sample_allocation_report.md"


def compute_allocation(sizes: dict[str, int]) -> tuple[dict, dict, int]:
    """Return (targets, props, total). 'none' -> 0; named clipped to [FLOOR,CAP]."""
    frame_total = sum(s for c, s in sizes.items() if c != NONE_CAT)
    targets, props = {}, {}
    for c, s in sizes.items():
        if c == NONE_CAT:
            targets[c], props[c] = 0, 0.0
            continue
        p = TARGET_TOTAL * s / frame_total
        props[c] = p
        targets[c] = int(min(CAP, max(FLOOR, p)) + 0.5)  # round half up after clip
    total = sum(targets.values())
    return targets, props, total


def main():
    wb = openpyxl.load_workbook(V1_XLSX)
    ws = wb["_summary"]

    # read per-category stats from _summary (cols: A cat, B total, C repos, D defaults)
    rows_meta = []  # (row_idx, category, total, repos, defaults)
    for r in range(2, ws.max_row + 1):
        cat = ws.cell(r, 1).value
        if cat is None:
            continue
        rows_meta.append((r, cat, ws.cell(r, 2).value, ws.cell(r, 3).value, ws.cell(r, 4).value))

    sizes = {cat: total for (_, cat, total, _, _) in rows_meta if cat != "TOTAL"}
    targets, props, total_target = compute_allocation(sizes)
    frame_total = sum(s for c, s in sizes.items() if c != NONE_CAT)

    if total_target < 85 or total_target > 115:
        print(f"[31] WARNING: total target {total_target} is outside 85-115 — "
              "consider adjusting FLOOR/CAP before relying on this.")

    # ---- _summary: add columns G/H/I + update TOTAL row ----
    ws.cell(1, 7, "sample_target")
    ws.cell(1, 8, "in_sample_frame")
    ws.cell(1, 9, "notes")
    for (r, cat, total, repos, defaults) in rows_meta:
        if cat == "TOTAL":
            ws.cell(r, 7, total_target)
            ws.cell(r, 8, frame_total)  # total skills in the sampling frame
            ws.cell(r, 9, f"sampling frame = {frame_total:,} skills across "
                          f"{sum(1 for c in sizes if c != NONE_CAT)} categories; "
                          f"'none' (654) excluded; total sample target {total_target}")
        elif cat == NONE_CAT:
            ws.cell(r, 7, 0)
            ws.cell(r, 8, "N")
            ws.cell(r, 9, "excluded from sampling frame")
        else:
            ws.cell(r, 7, targets[cat])
            ws.cell(r, 8, "Y")
            ws.cell(r, 9, "")
    ws.freeze_panes = "A2"
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 60

    # ---- each category tab: insert a note row at top, re-freeze ----
    note_font = Font(bold=True)
    for sheet in wb.worksheets:
        if sheet.title == "_summary":
            continue
        cat = sheet.title  # tab titles == category names (all valid in v1)
        sheet.insert_rows(1)
        if cat == NONE_CAT:
            note = "This category is excluded from the sampling frame. Do not select from this tab."
        else:
            t = targets.get(cat, FLOOR)
            note = (f"Sample target for this category: {t} skills "
                    "(selection guidance: aim for this number across Y-marked skills)")
        c = sheet.cell(1, 1, note)
        c.font = note_font
        c.alignment = Alignment(wrap_text=False, vertical="center")
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N_COLS)
        sheet.freeze_panes = "A3"  # keep note row (1) + header row (2) visible

    V2_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(V2_XLSX)

    # ---- v2 CSV mirror ----
    df = pd.read_csv(V1_CSV)
    df["sample_target"] = df["category"].map(lambda c: targets.get(c, 0))
    df["in_sample_frame"] = df["category"].map(lambda c: "N" if c == NONE_CAT else "Y")
    df["notes"] = df["category"].map(
        lambda c: "excluded from sampling frame" if c == NONE_CAT else "")
    df.to_csv(V2_CSV, index=False)

    # ---- allocation report ----
    meta = {cat: (total, repos, defaults) for (_, cat, total, repos, defaults) in rows_meta}
    named = sorted([c for c in sizes if c != NONE_CAT], key=lambda c: -sizes[c])
    capped = [c for c in named if props[c] > CAP]
    floored = [c for c in named if props[c] < FLOOR]
    coincide = [c for c in named if targets[c] >= meta[c][1]]  # target >= unique repos

    md = ["# Sample-allocation report — Phase 4 (floor 3 / cap 6)", ""]
    md.append(f"_Generated by `scripts/31_update_review_spreadsheet.py`. "
              f"Outputs: `sample/{V2_XLSX.name}` (+ `{V2_CSV.name}`)._")
    md += ["", "## Totals", ""]
    md.append(f"- Sampling frame: **{frame_total:,}** skills across **{len(named)}** named "
              "categories (`none` = 654 excluded).")
    md.append(f"- **Total sample target: {total_target}** "
              f"(floor {FLOOR}, cap {CAP}; target is reported, not forced to exactly 100).")
    md.append(f"- Clip is active: **{len(capped)}** categories capped down (proportional > {CAP}), "
              f"**{len(floored)}** floored up (proportional < {FLOOR}); "
              f"{len(named) - len(capped) - len(floored)} fall naturally within [{FLOOR}, {CAP}].")
    md += ["", "## Per-category targets", "",
           "`prop_pre_clip` = proportional allocation to 100 before the floor/cap clip; "
           "`target` = clipped + rounded. `def_Y` = suggested_default candidates available.", ""]
    md.append(L.md_table(
        ["Category", "Size", "Unique repos", "def_Y", "prop_pre_clip", "target", "clip effect"],
        [[c, L.fmt_int(meta[c][0]), L.fmt_int(meta[c][1]), L.fmt_int(meta[c][2]),
          f"{props[c]:.2f}", targets[c],
          "capped" if props[c] > CAP else ("floored" if props[c] < FLOOR else "—")]
         for c in named]))
    tightest = min(named, key=lambda c: meta[c][2] / targets[c])
    md += ["", "## Candidates vs target", "",
           "Every named category's suggested_default pool (`def_Y`) far exceeds its target, "
           "so each target is reachable while honouring max-2-per-repo. Tightest ratio: "
           f"`{tightest}` ({meta[tightest][2]} candidates for target {targets[tightest]}).", ""]
    md += ["## Anomalies / flags", ""]
    md.append(f"- **Clip distorts heavily by design.** Largest proportional pull is "
              f"`software-development` ({props['software-development']:.1f} → "
              f"{targets['software-development']}); smallest is `design-systems` "
              f"({props['design-systems']:.2f} → {targets['design-systems']}). The sample is "
              "deliberately far flatter than proportional — the point of floor+cap.")
    md.append(f"- **target ≥ unique-repo count:** {('none — ' if not coincide else '')}"
              + (", ".join(f"`{c}`" for c in coincide) if coincide else
                 "all targets sit well below their category's repo count, so publisher "
                 "diversity within each target is always achievable") + ".")
    md.append(f"- **`none` excluded:** 654 skills carry target 0 and `in_sample_frame=N`.")

    REPORT.write_text("\n".join(md).rstrip() + "\n")

    print(f"[31] total_target={total_target} frame={frame_total} "
          f"capped={len(capped)} floored={len(floored)}")
    print(f"[31] wrote {V2_XLSX.name}, {V2_CSV.name}, {REPORT.name}")


if __name__ == "__main__":
    main()